from flask import Flask, render_template, request, jsonify
import openpyxl
import os

app = Flask(__name__)

EXCEL_FILE = "library_records.xlsx"

# Ensure Excel file exists
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(["Accession No", "Author", "Title", "Call Number", "Location"])
        wb.save(EXCEL_FILE)

# Read records from Excel
def read_records():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # skip empty rows
            records.append({
                "accessionNo": row[0],
                "author": row[1],
                "title": row[2],
                "callNumber": row[3],
                "location": row[4]
            })
    return records

# Write a new record to Excel
def add_record(record):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        record["accessionNo"],
        record["author"],
        record["title"],
        record["callNumber"],
        record["location"]
    ])
    wb.save(EXCEL_FILE)

# Delete a record from Excel by accession number
def delete_record(accession_no):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == accession_no:
            ws.delete_rows(row)
            wb.save(EXCEL_FILE)
            return True
    return False


@app.route("/")
def home():
    return render_template("index.html")  # Your HTML file


@app.route("/records", methods=["GET"])
def get_records():
    return jsonify(read_records())


@app.route("/records", methods=["POST"])
def add_new_record():
    record = request.json
    # Check for duplicates
    existing = read_records()
    if any(r["accessionNo"] == record["accessionNo"] for r in existing):
        return jsonify({"error": "Accession number already exists!"}), 400
    add_record(record)
    return jsonify({"message": "Record added successfully!"})


@app.route("/records/<accession_no>", methods=["DELETE"])
def remove_record(accession_no):
    if delete_record(accession_no):
        return jsonify({"message": "Record deleted successfully!"})
    else:
        return jsonify({"error": "Record not found"}), 404

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
